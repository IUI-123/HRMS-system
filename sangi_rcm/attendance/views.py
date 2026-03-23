import json
import math
import openpyxl
from datetime import datetime, timedelta, time
from openpyxl.styles import PatternFill, Font
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.template.loader import get_template
from django.utils import timezone
from django.db.models import Count, Q
from xhtml2pdf import pisa
import csv
import calendar
import os
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    EmployeeProfile, Attendance, LeaveRequest, CompanyHoliday,
    CompanyNotice, DependentDetail, BackgroundVerification,
    SalaryUpdateRequest, ResignationRequest, PayrollUpdateRequest,
    CompanySettings, EmployeeDocument, Notification, AllowedIP, Task,
    Grievance  
)


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371000.0
    phi_1 = math.radians(lat1)
    phi_2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = math.sin(delta_phi / 2.0)**2 + math.cos(phi_1) * math.cos(phi_2) * math.sin(delta_lambda / 2.0)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c  

@never_cache
def login_view(request):
    if request.method == 'POST':
        u_name = request.POST.get('username')
        p_word = request.POST.get('password')
        user = authenticate(request, username=u_name, password=p_word)
        
        if user is not None:
            login(request, user)
            try:
                profile = user.employeeprofile
                
                # 🟢 THE SMART ROUTER
                if profile.is_hr:
                    return redirect('hr_dashboard')
                elif getattr(profile, 'is_accounts', False):
                    return redirect('accounts_dashboard') 
                elif profile.is_hod:  # 🟢 NEW: HOD Redirect
                    return redirect('hod_dashboard')
                else:
                    return redirect('employee_dashboard')
                    
            except EmployeeProfile.DoesNotExist:
                return redirect('/admin/')
        else:
            messages.error(request, '❌ Invalid username or password.')
            return redirect('login_view')
            
    return render(request, 'login.html')

@never_cache
def logout_view(request):
    logout(request)
    return redirect('login_view')

@login_required
@never_cache
def employee_dashboard(request):
    # If the user doesn't have a profile (like a Superuser), send them to the admin panel!
    if not hasattr(request.user, 'employeeprofile'):
        if request.user.is_superuser:
            return redirect('/admin/')
        else:
            messages.error(request, "Your account has no employee profile assigned. Contact Admin.")
            return redirect('login') 

    # 1. Setup Profile
    profile = request.user.employeeprofile
    today = timezone.localtime(timezone.now()).date()

    # 🟢 NEW: CATCH THE BGV FORM SUBMISSION (AJAX)
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'submit_bgv':
            # 1. Save Background/Experience Details
            bgv, created = BackgroundVerification.objects.get_or_create(employee=profile)
            bgv.ref_name = request.POST.get('ref_name')
            bgv.ref_post = request.POST.get('ref_post')
            bgv.ref_organization = request.POST.get('ref_organization')
            bgv.ref_phone = request.POST.get('ref_phone')
            bgv.save()

            # 2. Save Dependents (We delete old ones and recreate to avoid duplicates if they edit)
            DependentDetail.objects.filter(employee=profile).delete()
            for i in range(1, 4): # Loop for 3 possible dependents
                dep_name = request.POST.get(f'dep_name_{i}')
                dep_relation = request.POST.get(f'dep_relation_{i}')
                dep_dob = request.POST.get(f'dep_dob_{i}')
                
                if dep_name and dep_relation and dep_dob:
                    DependentDetail.objects.create(
                        employee=profile, name=dep_name, relation=dep_relation, dob=dep_dob
                    )
            
            # 🟢 GUARANTEED JSON RESPONSE FOR JS
            return JsonResponse({'success': True, 'message': '✅ BGV & Dependent Details Saved!'})


    try:
        current_month = int(request.GET.get('month', today.month))
        current_year = int(request.GET.get('year', today.year))
    except ValueError:
        current_month = today.month
        current_year = today.year

    # 2. Basic Attendance & Leave Data
    my_attendance = Attendance.objects.filter(employee=profile)
    monthly_attendance = my_attendance.filter(date__year=current_year, date__month=current_month)
    my_leaves = LeaveRequest.objects.filter(employee=profile).order_by('-start_date')

    # 3. Calendar Logic (Preparing JSON for the frontend)
    cal_data = {}
    for a in my_attendance:
        cal_data[str(a.date)] = a.status 
        
    for leave in my_leaves.filter(status='Approved'):
        delta = leave.end_date - leave.start_date
        for i in range(delta.days + 1):
            leave_day = leave.start_date + timedelta(days=i)
            cal_data[str(leave_day)] = 'Leave' 
            
    # 4. Dashboard Stats
    today_attendance = my_attendance.filter(date=today).first()
    present_count = my_attendance.filter(status='P').count()
    late_count = my_attendance.filter(status='L').count()
    absent_count = my_attendance.filter(status='A').count()
    half_day_count = my_attendance.filter(status='H').count() 
    early_leave_count = my_attendance.filter(status='E').count() 

    # Dynamic Monthly Stats
    total_present = monthly_attendance.filter(status='P').count()
    total_absent = monthly_attendance.filter(status='A').count()
    total_late = monthly_attendance.filter(status='L').count()
    total_leave = monthly_attendance.filter(status='V').count() + \
                  LeaveRequest.objects.filter(
                      employee=profile,
                      start_date__year=current_year,
                      start_date__month=current_month,
                      status='Approved'
                  ).count()

    # 🟢 THE SMART NOTICE FILTER
    my_notices = CompanyNotice.objects.filter(
        Q(target_audience='All') |
        Q(target_audience='Department', target_department=profile.department) |
        Q(target_audience='Employee', target_employee=request.user)
    ).order_by('-created_at')[:5]
    
    notice_count = my_notices.count()
    leave_count = my_leaves.filter(status='Approved').count() 
    upcoming_holidays = CompanyHoliday.objects.filter(date__gte=today).order_by('date')
    my_resignations = ResignationRequest.objects.filter(employee=profile).order_by('-applied_date')

    # 🟢 NEW: FETCH EXISTING BGV DATA TO SHOW ON THE FORM
    existing_bgv = BackgroundVerification.objects.filter(employee=profile).first()
    existing_dependents = DependentDetail.objects.filter(employee=profile)

    # Grab dependents and safely split them into 1, 2, and 3
    deps = list(DependentDetail.objects.filter(employee=profile))
    dep_1 = deps[0] if len(deps) > 0 else None
    dep_2 = deps[1] if len(deps) > 1 else None
    dep_3 = deps[2] if len(deps) > 2 else None

    # 5. Context Construction
    context = {
        'profile': profile,
        'my_leaves': my_leaves,
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
        'half_day_count': half_day_count,
        'early_leave_count': early_leave_count, 
        'leave_count': leave_count,
        'calendar_data': json.dumps(cal_data),
        'holidays': upcoming_holidays,
        'today_attendance': today_attendance, 
        'my_resignations': my_resignations,
        'my_notices': my_notices,
        'notice_count': notice_count,
        'existing_bgv': existing_bgv,                  # 🟢 Sent to HTML
        'existing_dependents': existing_dependents,    # 🟢 Sent to HTML
        'existing_bgv': existing_bgv,
        'dep_1': dep_1,
        'dep_2': dep_2,
        'dep_3': dep_3,
        'attendance_logs': monthly_attendance.order_by('-date'),
        'total_present': total_present,
        'total_absent': total_absent,
        'total_late': total_late,
        'total_leave': total_leave,
        'current_month': current_month,
        'current_year': current_year,
    }

    return render(request, 'employee_dashboard.html', context)

@login_required
@never_cache
def hr_dashboard(request):
    # 1. Access Control
    if not getattr(request.user.employeeprofile, 'is_hr', False):
        return redirect('employee_dashboard')
        
    # 2. THE POST HANDLER
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # 🟢 THE NEW AJAX ADD EMPLOYEE BLOCK
        if action == 'add_employee':
            return add_employee(request)

        # 🟢 THE SMART DOCUMENT SAVER
        if action == 'upload_documents':
            emp_id = request.POST.get('employee_id')
            emp = get_object_or_404(EmployeeProfile, id=emp_id)
            
            doc_fields = [
                'resume_cv', 'photo', 'pan_card_doc', 'tenth_mark', 'twelfth_mark', 
                'graduation', 'experience_letter', 'aadhar_card', 'passbook'
            ]
                    
            files_uploaded = 0
            for field in doc_fields:
                if field in request.FILES:
                    setattr(emp, field, request.FILES[field])
                    files_uploaded += 1
            
            if files_uploaded > 0:
                emp.save()
                msg = f"Vault Updated! {files_uploaded} document(s) saved."
                success_status = True
            else:
                msg = "No files were selected to upload."
                success_status = False

            # AJAX RESPONSE: No page refresh!
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': success_status, 'message': msg})
                
            if success_status:
                messages.success(request, msg)
            else:
                messages.warning(request, msg)
            return redirect('hr_dashboard')

    # 3. General HR Dashboard Data (Live Monitoring & Staff)
    today = timezone.localtime(timezone.now()).date()
    todays_attendance = Attendance.objects.filter(date=today).select_related('employee__user')
    pending_leaves = LeaveRequest.objects.filter(status='Pending').select_related('employee__user')
    all_resignations = ResignationRequest.objects.select_related('employee__user').all().order_by('-applied_date')
    
    # 🟢 ONLY ACTIVE EMPLOYEES IN THE TABLE
    employees = EmployeeProfile.objects.select_related('user').filter(is_hr=False, user__is_active=True)
    total_employees = employees.count()
    
    # 🟢 MONTHLY FILTERING FOR COMPANY-WIDE STATS
    try:
        current_month = int(request.GET.get('month', today.month))
        current_year = int(request.GET.get('year', today.year))
    except ValueError:
        current_month = today.month
        current_year = today.year
    
    # ⚠️ STEP A: BASE QUERY (COMPANY-WIDE MONTHLY)
    monthly_attendance_base = Attendance.objects.filter(
        date__year=current_year, 
        date__month=current_month
    )
    
    # ⚠️ STEP B: COMPANY-WIDE MONTHLY MATH
    total_present = monthly_attendance_base.filter(status='P').count()
    total_absent = monthly_attendance_base.filter(status='A').count()
    total_late = monthly_attendance_base.filter(status='L').count()
    
    # ⚠️ STEP C: SLICE IT FOR THE TABLE
    monthly_attendance = monthly_attendance_base.select_related('employee__user').order_by('-date')[:200]
    
    # Company-wide monthly leaves
    total_leave = LeaveRequest.objects.filter(
        status='Approved',
        start_date__year=current_year,
        start_date__month=current_month
    ).count()
    
    # 🟢 NEW: COMPANY-WIDE STATS FOR *TODAY* (Live Monitoring Cards)
    present_today = todays_attendance.filter(status__in=['P', 'L', 'H', 'E']).count()
    late_today = todays_attendance.filter(status='L').count()
    half_day_today = todays_attendance.filter(status='H').count()
    early_leave_today = todays_attendance.filter(status='E').count()
    absent_today = todays_attendance.filter(status='A').count()

    # ==========================================
    # 4. HR's PERSONAL ATTENDANCE DATA ("My Attendance" Tab)
    # ==========================================
    my_profile = request.user.employeeprofile
    
    # ⚠️ Filter personal attendance by the selected month/year!
    my_attendance = Attendance.objects.filter(
        employee=my_profile,
        date__year=current_year,
        date__month=current_month
    ).order_by('-date')
    
    my_today_attendance = Attendance.objects.filter(employee=my_profile, date=today).first() 
    
    my_leave_history = LeaveRequest.objects.filter(employee=my_profile).order_by('-start_date')
    my_approved_leaves = my_leave_history.filter(
        status='Approved',
        start_date__year=current_year,
        start_date__month=current_month
    )

    # 🟢 ONLY ACTIVE EMPLOYEES IN THE DROPDOWN
    all_employees = EmployeeProfile.objects.select_related('user').filter(user__is_active=True).order_by('department', 'user__first_name')
    
    cal_data = {}
    for a in my_attendance:
        cal_data[str(a.date)] = a.status 
        
    for leave in my_approved_leaves:
        delta = leave.end_date - leave.start_date
        for i in range(delta.days + 1):
            leave_day = leave.start_date + timedelta(days=i)
            cal_data[str(leave_day)] = 'Leave' 
            
    # HR's Personal Monthly Math
    my_present_count = my_attendance.filter(status='P').count()
    my_late_count = my_attendance.filter(status='L').count()
    my_absent_count = my_attendance.filter(status='A').count()
    my_leave_count = my_approved_leaves.count()
    
    upcoming_holidays = CompanyHoliday.objects.filter(date__gte=today).order_by('date')
    context = {
        'employees': employees,
        'total_employees': total_employees,
        'todays_attendance': todays_attendance,
        'pending_leaves': pending_leaves,
        'all_resignations': all_resignations,
        'present_today': present_today,
        'late_today': late_today,
        'half_day_today': half_day_today,
        'early_leave_today': early_leave_today,
        'absent_today': absent_today,
        'my_today_attendance': my_today_attendance,
        'my_leave_history': my_leave_history,
        'cal_data': cal_data,
        'my_present_count': my_present_count,
        'my_late_count': my_late_count,
        'my_absent_count': my_absent_count,
        'my_leave_count': my_leave_count,
        'upcoming_holidays': upcoming_holidays,
        'all_employees': all_employees, 
        'attendance_logs': monthly_attendance,
        'total_present': total_present,
        'total_absent': total_absent,
        'total_late': total_late,
        'total_leave': total_leave,
        'current_month': current_month,
        'current_year': current_year,
        'my_attendance': my_attendance,  # ADD THIS LINE!
        'attendance_logs': monthly_attendance,
    }
    return render(request, 'hr_dashboard.html', context)

@login_required
def apply_leave(request):
    if request.method == 'POST':
        LeaveRequest.objects.create(
            employee=request.user.employeeprofile,
            subject=request.POST.get('subject'),
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date'),
            description=request.POST.get('description')
        )
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def update_leave(request, leave_id, action):
    if getattr(request.user.employeeprofile, 'is_hr', False):
        if request.method == 'POST':
            try:
                leave = LeaveRequest.objects.get(id=leave_id)
                if action in ['Approved', 'Rejected']:
                    leave.status = action
                    leave.save()
            except LeaveRequest.DoesNotExist:
                pass
    return redirect('hr_dashboard')


@login_required
def mark_attendance(request):
    if request.method == 'POST':
        user_ip = get_client_ip(request)
        
        # 🟢 DYNAMIC SECURITY SHIELD
        # Grabs all the IPs you saved in the database
        allowed_ips = AllowedIP.objects.values_list('ip_address', flat=True)
        
        # If their IP isn't in the database list, block them!
        if user_ip not in allowed_ips:
            return JsonResponse({
                "success": False, 
                "message": f"❌ Punch rejected! You must be on the Office Network. (Your IP: {user_ip})"
            })
            
        # ... (keep the rest of your attendance saving logic below this) ...
            
        try:
            user_lat = float(request.POST.get('latitude'))
            user_lon = float(request.POST.get('longitude'))
            OFFICE_LAT = 28.587196 
            OFFICE_LON = 77.315562
            
            distance = calculate_distance(user_lat, user_lon, OFFICE_LAT, OFFICE_LON)

            # 🛑 SECURITY SHIELD ACTIVATED
            if distance > 45:
                return JsonResponse({"success": False, "message": f"❌ Punch rejected! You are {int(distance)} meters away."})
        except (TypeError, ValueError):
            return JsonResponse({"success": False, "message": "❌ GPS Location not found!"})

        employee = request.user.employeeprofile
        
        # 🟢 NOTE: Facial recognition system removed post-ZKTeco migration
        # Attendance now handled by biometric machine hardware integration (zkteco_cdata/zkteco_getrequest)
        # Legacy API kept for backward compatibility with manual punch entry
        
        # ==========================================
        # 🟢 THE TIME MATRIX ENGINE (Timezone Fixed)
        # ==========================================
        now = timezone.localtime(timezone.now())  # Grabs the exact local time from settings.py!
        today = now.date()
        current_time = now.time()
        
        # Check if they already have an attendance record for today
        attendance = Attendance.objects.filter(employee=employee, date=today).first()
        
   # 🌅 PUNCH IN LOGIC (Morning)
        if not attendance:
            if current_time > time(14, 0): morning_status = 'A'
            elif current_time > time(13, 30): morning_status = 'H'
            elif current_time > time(9, 30): morning_status = 'L'
            else: morning_status = 'P'
                
            Attendance.objects.create(
                employee=employee,
                date=today,
                check_in=current_time,  # 🟢 FIXED
                status=morning_status,
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),
                selfie=selfie_file,
                ip_address=user_ip
            )
            return JsonResponse({"success": True, "message": f"✅ Identity Verified & Punched In! (Status: {morning_status})"})
            
        # 🌇 PUNCH OUT LOGIC (Evening)
        else:
            if attendance.check_out:
                return JsonResponse({"success": False, "message": "You have already punched out for today!"})
                
            attendance.check_out = current_time
            attendance.selfie = selfie_file      # 🟢 FIXED: It now saves the Evening Selfie!
            in_time = attendance.check_in
            out_time = current_time
            
            # 🟢 FIXED: Removed the 'if not is_hr_edited' block so it ALWAYS recalculates the stopwatch!
            if in_time > time(14, 0): morning_status = 'A'
            elif in_time > time(13, 30): morning_status = 'H'
            elif in_time > time(9, 30): morning_status = 'L'
            else: morning_status = 'P'

            if out_time < time(13, 30): evening_status = 'A'
            elif out_time < time(16, 0): evening_status = 'H'
            elif out_time < time(17, 30): evening_status = 'E'
            else: evening_status = 'P'

            if morning_status == 'A' or evening_status == 'A': final_status = 'A'
            elif morning_status == 'H' or evening_status == 'H': final_status = 'H'
            elif morning_status == 'L' and evening_status == 'E': final_status = 'H'
            elif morning_status == 'L': final_status = 'L'
            elif evening_status == 'E': final_status = 'E'
            else: final_status = 'P'
                
            # ⏱️ THE STOPWATCH OVERRIDE (Total Hours Check)
            in_dt = datetime.combine(today, in_time)
            out_dt = datetime.combine(today, out_time)
            duration_hours = (out_dt - in_dt).total_seconds() / 3600.0
            
            if duration_hours < 4.5:
                final_status = 'A'  # Less than 4.5 hours = Automatic Absent
            elif duration_hours < 8.5 and final_status not in ['A', 'H']:
                final_status = 'H'  # Less than 8.5 hours = Automatic Half Day

            attendance.status = final_status
            attendance.save()
            
            return JsonResponse({"success": True, "message": f"✅ Identity Verified & Punched Out! Final Status: {attendance.status}"})

    return JsonResponse({"success": False, "message": "Invalid request"})
# 🛑 BULLETPROOF ADD EMPLOYEE
@login_required
@never_cache
def add_employee(request):
    if request.method == 'POST' and getattr(request.user.employeeprofile, 'is_hr', False):
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        emp_id = request.POST.get('emp_id')
        dept = request.POST.get('department')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # 🟢 Extract the HOD role from the AJAX form data
        is_hod_str = request.POST.get('is_hod', 'False')
        is_hod_boolean = True if is_hod_str == 'True' else False

        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'message': 'Username already taken.'})
        if EmployeeProfile.objects.filter(emp_id=emp_id).exists():
            return JsonResponse({'success': False, 'message': 'Employee ID already exists.'})

        user = User.objects.create_user(username=username, password=password, first_name=first_name, last_name=last_name)
        
        # 🟢 WE CAPTURE AND SAVE THE HOD FLAG HERE
        new_emp = EmployeeProfile.objects.create(
            user=user, 
            emp_id=emp_id, 
            department=dept,
            is_hod=is_hod_boolean  # <-- Saves to the database
        )
        
        # 🟢 RETURN THE DATA FOR THE JAVASCRIPT TABLE
        return JsonResponse({
            'success': True, 
            'message': f'{first_name} added successfully!',
            'employee': {
                'db_id': new_emp.id,               
                'emp_id': new_emp.emp_id,          
                'first_name': new_emp.user.first_name,
                'last_name': new_emp.user.last_name,
                'department': new_emp.department,
                'is_hod': new_emp.is_hod  # <-- Sent back to UI just in case!
            }
        })
    return JsonResponse({'success': False, 'message': 'Invalid Request'})

@login_required
def hr_edit_attendance(request):
    if request.method == 'POST':
        attendance_id = request.POST.get('attendance_id')
        new_status = request.POST.get('new_status')
        
        # 🟢 THE BUG WAS HERE! It must match 'edit_reason' from the HTML form.
        reason = request.POST.get('edit_reason') 

        attendance = get_object_or_404(Attendance, id=attendance_id)
        attendance.status = new_status
        attendance.is_hr_edited = True  # 🟢 Security switch ON
        
        # 🟢 Saves the exact reason to the database
        attendance.employee_note = f"HR Override: {reason}" 
        attendance.save()

        messages.success(request, f"Attendance successfully overridden!")
        
    return redirect('hr_dashboard') # Or wherever you want it to redirect

@login_required
def export_attendance(request):
    # Security Check
    profile = getattr(request.user, 'employeeprofile', None)
    if not profile or (not profile.is_hr and not profile.is_accounts):
        return HttpResponse("Unauthorized", status=401)

    # Setup the exact Excel response format (.xlsx)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Sangi_Attendance_{timezone.localtime(timezone.now()).strftime("%b_%Y")}.xlsx"'

    # Create the Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    # 🎨 DEFINE OUR STYLES
    header_fill = PatternFill(start_color="0B1A3A", end_color="0B1A3A", fill_type="solid") # Dark Blue
    col_fill = PatternFill(start_color="2A62C9", end_color="2A62C9", fill_type="solid")   # Table Header Blue
    white_bold = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(color="FFFFFF", bold=True, size=16)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin', color="CCCCCC"), right=Side(style='thin', color="CCCCCC"), 
                         top=Side(style='thin', color="CCCCCC"), bottom=Side(style='thin', color="CCCCCC"))

    # 🟢 1. CREATE THE BIG HEADER
    ws.merge_cells('A1:L2')
    title_cell = ws['A1']
    title_cell.value = f"🏢 MONTHLY ATTENDANCE REGISTER — {timezone.now().strftime('%b %Y')}"
    title_cell.fill = header_fill
    title_cell.font = title_font
    title_cell.alignment = left_align

    # 🟢 2. CREATE THE COLUMN HEADERS
    columns = ["Emp ID", "Name", "Department", "Date", "Day", "In Time", "Out Time", "Hours Worked", "Status", "HR Edited?", "Edit Reason", "Employee Note"]
    
    for col_num, col_name in enumerate(columns, 1):
        c = ws.cell(row=3, column=col_num, value=col_name)
        c.fill = col_fill
        c.font = white_bold
        c.alignment = center_align
        c.border = thin_border

    # Set Column Widths for clean layout
    widths = [10, 22, 15, 15, 12, 12, 12, 15, 18, 12, 25, 25]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 🟢 3. FETCH AND SORT DATA
    records = Attendance.objects.all().select_related('employee', 'employee__user').order_by('employee__emp_id', '-date')

    # Status Colors Dictionary
    status_colors = {
        'P': ("16A34A", "FFFFFF"), # Green
        'A': ("DC2626", "FFFFFF"), # Red
        'L': ("EA580C", "FFFFFF"), # Orange
        'H': ("7E22CE", "FFFFFF"), # Purple
        'E': ("D97706", "FFFFFF"), # Yellow/Orange
        'W': ("2563EB", "FFFFFF"), # Blue
        'O': ("B45309", "FFFFFF"), # Brown
    }

    row_num = 4
    current_emp_id = None

    for record in records:
        # Add visual separation block for new employees
        if current_emp_id is not None and current_emp_id != record.employee.emp_id:
            ws.merge_cells(f'A{row_num}:L{row_num}')
            ws.cell(row=row_num, column=1).fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            row_num += 1

        # Calculate Hours Worked
        hours_worked = "—"
        if record.check_in and record.check_out:
            # 🟢 FIXED: We now use 'record.date' instead of 'date.today()'
            # This makes the math perfectly accurate for historical reports!
            in_dt = datetime.combine(record.date, record.check_in)
            out_dt = datetime.combine(record.date, record.check_out)
            diff = out_dt - in_dt
            
            hrs = diff.seconds // 3600
            mins = (diff.seconds % 3600) // 60
            hours_worked = f"{hrs}h {mins}m"

        # Determine Department (Fallback to 'IT' if not set in model)
        dept = getattr(record.employee, 'department', 'IT')

        row_data = [
            record.employee.emp_id,
            f"{record.employee.user.first_name} {record.employee.user.last_name}",
            dept,
            record.date.strftime('%d %b %Y'),
            record.date.strftime('%A'),
            record.check_in.strftime('%H:%M') if record.check_in else "—",
            record.check_out.strftime('%H:%M') if record.check_out else "—",
            hours_worked,
            record.get_status_display(),
            "YES" if record.is_hr_edited else "No",
            record.employee_note.replace("HR Override: ", "").replace("Manual: ", "") if record.employee_note else "", 
            "" # Spare column for extra notes
        ]

        # 🟢 4. WRITE AND PAINT THE ROWS
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.border = thin_border
            c.alignment = center_align

            # Color the Status Badge!
            if col_idx == 9 and record.status in status_colors:
                bg_color, fg_color = status_colors[record.status]
                c.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                c.font = Font(color=fg_color, bold=True)
            
            # Highlight HR Edits!
            if col_idx == 10 and val == "YES":
                c.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # Yellow highlight
                c.font = Font(color="854D0E", bold=True) # Dark gold text

        current_emp_id = record.employee.emp_id
        row_num += 1

    wb.save(response)
    return response
@login_required
def request_missed_swap(request):
    if request.method == 'POST':
        reason = request.POST.get('reason')
        today = timezone.localtime(timezone.now()).date()
        current_time = datetime.now().time()
        
        Attendance.objects.update_or_create(
            employee=request.user.employeeprofile,
            date=today,
            defaults={
                'status': 'M',
                'employee_note': reason,
                'check_in': current_time 
            }
        )
    return redirect('employee_dashboard')

@login_required
def hr_manual_entry(request):
    if request.method == 'POST':
        emp_id = request.POST.get('employee_id')
        in_time_str = request.POST.get('in_time')
        out_time_str = request.POST.get('out_time')
        reason = request.POST.get('reason')

        employee = get_object_or_404(EmployeeProfile, emp_id=emp_id)
        today = timezone.localtime(timezone.now()).date()
        
        # Convert HTML time strings to Python Time objects
        in_time = datetime.strptime(in_time_str, '%H:%M').time() if in_time_str else None
        out_time = datetime.strptime(out_time_str, '%H:%M').time() if out_time_str else None

        # 🧠 1. Calculate Morning Status
        morning_status = 'P'
        if in_time:
            if in_time > time(14, 0): morning_status = 'A'
            elif in_time > time(13, 30): morning_status = 'H'
            elif in_time > time(9, 30): morning_status = 'L'
            else: morning_status = 'P'

        # 🧠 2. Calculate Evening Status
        evening_status = 'P'
        if out_time:
            if out_time < time(13, 30): evening_status = 'A'
            elif out_time < time(16, 0): evening_status = 'H'
            elif out_time < time(17, 30): evening_status = 'E'
            else: evening_status = 'P'

        # 🧠 3. The Ultimate Matrix Combo
        final_status = 'P'
        if in_time and not out_time:
            final_status = morning_status 
        elif in_time and out_time:
            if morning_status == 'A' or evening_status == 'A': final_status = 'A'
            elif morning_status == 'H' or evening_status == 'H': final_status = 'H'
            elif morning_status == 'L' and evening_status == 'E': final_status = 'H'
            elif morning_status == 'L': final_status = 'L'
            elif evening_status == 'E': final_status = 'E'
            else: final_status = 'P'
            
            # ⏱️ THE STOPWATCH OVERRIDE (Total Hours Check)
            in_dt = datetime.combine(today, in_time)
            out_dt = datetime.combine(today, out_time)
            duration_hours = (out_dt - in_dt).total_seconds() / 3600.0
            
            if duration_hours < 4.5:
                final_status = 'A'  # Less than 4.5 hours = Automatic Absent
            elif duration_hours < 8.5 and final_status not in ['A', 'H']:
                final_status = 'H'  # Less than 8.5 hours = Automatic Half Day

        # Update or Create the Database Record!
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            date=today,
            defaults={
                'status': final_status, 
                'check_in': in_time, 
                'check_out': out_time, 
                'is_hr_edited': True,                        # 🟢 SECURITY SWITCH ON!
                'employee_note': f"Manual Entry: {reason}"   # 🟢 REASON SAVED!
            }
        )

        if not created:
            if in_time: attendance.check_in = in_time
            if out_time: attendance.check_out = out_time
            attendance.status = final_status
            attendance.is_hr_edited = True                   # 🟢 SECURITY SWITCH ON!
            attendance.employee_note = f"HR Override: {edit_reason}" 
            attendance.save()

        messages.success(request, f"Manual Entry successfully logged for {employee.user.first_name}!")
        return redirect('hr_dashboard')
# 🟢 THE NEW REMOVE EMPLOYEE FUNCTION
@login_required
def remove_employee(request, emp_id):
    if getattr(request.user.employeeprofile, 'is_hr', False):
        if request.method == 'POST':
            try:
                employee = EmployeeProfile.objects.get(emp_id=emp_id)
                
                # Disable the user account
                user = employee.user
                user.is_active = False
                user.save()
                
                # Disable the profile
                employee.is_active = False
                employee.save()
                    
            except EmployeeProfile.DoesNotExist:
                pass
                
    return redirect('hr_dashboard')


# 🟢 1. HR LEAVE APPROVAL LOGIC
@login_required
def hr_update_leave(request, leave_id):
    if request.method == 'POST' and getattr(request.user.employeeprofile, 'is_hr', False):
        leave = get_object_or_404(LeaveRequest, id=leave_id)
        action = request.POST.get('action')
        
        if action == 'approve_paid':
            leave.status = 'Approved'
            leave.is_paid = True
            msg = f"Leave for {leave.employee.user.first_name} Approved (Paid)."
        elif action == 'approve_unpaid':
            leave.status = 'Approved'
            leave.is_paid = False
            msg = f"Leave for {leave.employee.user.first_name} Approved (Unpaid)."
        elif action == 'reject':
            leave.status = 'Rejected'
            msg = f"Leave for {leave.employee.user.first_name} Rejected."
            
        leave.save()
        
        # If it's an AJAX request, return JSON so the page doesn't blink
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': msg})
            
        messages.success(request, msg)
    return redirect('hr_dashboard')


# 🟢 2. THE ACCOUNTS PAYROLL ENGINE
import json
import calendar
from datetime import datetime, date, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import JsonResponse
from .models import EmployeeProfile, Attendance, LeaveRequest, ResignationRequest, CompanyHoliday, BackgroundVerification, DependentDetail

@login_required
def accounts_dashboard(request):
    profile = getattr(request.user, 'employeeprofile', None)
    if not profile or not getattr(profile, 'is_accounts', False):
        return redirect('employee_dashboard')

    # ==========================================
    # 1. CATCH THE BGV FORM SUBMISSION (AJAX)
    # ==========================================
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'submit_bgv':
            bgv, created = BackgroundVerification.objects.get_or_create(employee=profile)
            bgv.ref_name = request.POST.get('ref_name')
            bgv.ref_post = request.POST.get('ref_post')
            bgv.ref_organization = request.POST.get('ref_organization')
            bgv.ref_phone = request.POST.get('ref_phone')
            bgv.save()

            DependentDetail.objects.filter(employee=profile).delete()
            for i in range(1, 4):
                dep_name = request.POST.get(f'dep_name_{i}')
                dep_relation = request.POST.get(f'dep_relation_{i}')
                dep_dob = request.POST.get(f'dep_dob_{i}')
                
                if dep_name and dep_relation and dep_dob:
                    DependentDetail.objects.create(employee=profile, name=dep_name, relation=dep_relation, dob=dep_dob)
            
            return JsonResponse({'success': True, 'message': '✅ BGV & Dependent Details Saved!'})

    # ==========================================
    # 2. SMART PAYROLL ENGINE (Month-Aware)
    # ==========================================
    now = timezone.localtime(timezone.now())
    try:
        current_month = int(request.GET.get('month', now.month))
        current_year = int(request.GET.get('year', now.year))
    except ValueError:
        current_month = now.month
        current_year = now.year

    # 🟢 DYNAMIC DAYS CALCULATION (Feb=28/29, Mar=31, etc.)
    # Using python's calendar module to get exact days in the selected month
    days_in_month = calendar.monthrange(current_year, current_month)[1]
    month_start = date(current_year, current_month, 1)
    month_end = date(current_year, current_month, days_in_month)

    # Exclude the Accounts Manager (profile) from this list if you only want to process staff
    employees = EmployeeProfile.objects.select_related('user').filter(user__is_active=True).exclude(id=profile.id)
    payroll_data = []
    
    upcoming_holidays = CompanyHoliday.objects.filter(date__gte=now.date()).order_by('date')

    for emp in employees:
        # ⚠️ STRICT MONTHLY FILTER: Fetch only this month's attendance for the employee
        records = Attendance.objects.filter(
            employee=emp, 
            date__year=current_year, 
            date__month=current_month
        )
        
        # Base Counts
        full_paid_days = records.filter(status__in=['P', 'W', 'O', 'E', 'L']).count() 
        half_days = records.filter(status='H').count()
        lates = records.filter(status='L').count()
        early_leaves = records.filter(status='E').count() 
        
        paid_leaves_count = 0
        unpaid_leaves_count = 0
        
        # 🟢 SMART OVERLAP LEAVE CALCULATION
        # Fetches leaves that touch this month, even if they started last month
        approved_leaves = LeaveRequest.objects.filter(
            employee=emp, 
            status='Approved',
            start_date__lte=month_end,
            end_date__gte=month_start
        )
        
        for leave in approved_leaves:
            # Calculate how many days of this leave fall strictly inside the current month
            overlap_start = max(leave.start_date, month_start)
            overlap_end = min(leave.end_date, month_end)
            days = (overlap_end - overlap_start).days + 1
            
            if days > 0:
                # Assuming your LeaveRequest model has an 'is_paid' boolean
                if getattr(leave, 'is_paid', False): 
                    paid_leaves_count += days
                else: 
                    unpaid_leaves_count += days

        # 🟢 PENALTY LOGIC
        # Every 2 lates = 0.5 day deduction. Every 2 early leaves = 0.5 day deduction.
        late_penalty_days = (lates // 2) * 0.5
        early_leave_penalty_days = (early_leaves // 2) * 0.5
        
        # Calculate Total Payable Days
        total_payable_days = float(full_paid_days + paid_leaves_count) + (half_days * 0.5) - late_penalty_days - early_leave_penalty_days
        
        # Floor check: you can't have negative payable days
        if total_payable_days < 0:
            total_payable_days = 0

        # 🟢 FINAL SALARY CALCULATION
        # Convert base_salary to float to avoid Decimal/Float math errors
        base_salary_float = float(emp.base_salary) if emp.base_salary else 0.0
        per_day_salary = base_salary_float / days_in_month if days_in_month > 0 else 0
        
        calculated_net_salary = float(total_payable_days) * per_day_salary
        emp_resignations = ResignationRequest.objects.filter(employee=emp).order_by('-applied_date')

        # Append calculated data to the list that gets sent to HTML
        payroll_data.append({
            'profile': emp,
            'presents': full_paid_days, 
            'lates': lates,             
            'paid_leaves': paid_leaves_count,
            'unpaid_leaves': unpaid_leaves_count,
            'payable_days': total_payable_days,
            'net_salary': round(calculated_net_salary, 2),
            'my_resignations': emp_resignations,
        })
# ==========================================
    # 3. ACCOUNTS MANAGER PERSONAL DATA
    # ==========================================
    my_attendance = Attendance.objects.filter(employee=profile)
    my_leaves = LeaveRequest.objects.filter(employee=profile).order_by('-start_date')
    my_today_attendance = my_attendance.filter(date=timezone.localtime(timezone.now()).date()).first()
    my_resignations = ResignationRequest.objects.filter(employee=profile).order_by('-applied_date')
    
    # 🟢 COMPANY-WIDE MONTHLY ATTENDANCE STATS
    
    # STEP A: The Base Query
    monthly_all_attendance_base = Attendance.objects.filter(
        date__year=current_year,
        date__month=current_month
    )
    
    # STEP B: Do the math
    total_present = monthly_all_attendance_base.filter(status='P').count()
    total_absent = monthly_all_attendance_base.filter(status='A').count()
    total_late = monthly_all_attendance_base.filter(status='L').count()
    
    # STEP C: Slice it for the table so the UI doesn't lag
    attendance_logs = monthly_all_attendance_base.order_by('-date')[:200]
    
    monthly_all_leaves = LeaveRequest.objects.filter(
        status='Approved',
        start_date__year=current_year,
        start_date__month=current_month
    )
    total_leave = monthly_all_leaves.count()
    
    cal_data = {}
    for a in my_attendance:
        cal_data[str(a.date)] = a.status 
        
    for leave in my_leaves.filter(status='Approved'):
        delta = leave.end_date - leave.start_date
        for i in range(delta.days + 1):
            leave_day = leave.start_date + timedelta(days=i)
            cal_data[str(leave_day)] = 'Leave' 
            
    my_present_count = my_attendance.filter(status='P').count()
    my_late_count = my_attendance.filter(status='L').count()
    my_absent_count = my_attendance.filter(status='A').count()
    my_leave_count = my_leaves.filter(status='Approved').count()

    existing_bgv = BackgroundVerification.objects.filter(employee=profile).first()
    deps = list(DependentDetail.objects.filter(employee=profile))
    dep_1 = deps[0] if len(deps) > 0 else None
    dep_2 = deps[1] if len(deps) > 1 else None
    dep_3 = deps[2] if len(deps) > 2 else None

    context = {
        'month': current_month,
        'year': current_year,
        'month_name': calendar.month_name[current_month],
        'payroll_data': payroll_data, # Assuming you define this above in your view
        'all_employees': employees, # Assuming you define this above in your view
        'my_present_count': my_present_count,
        'my_late_count': my_late_count,
        'my_absent_count': my_absent_count,
        'my_leave_count': my_leave_count,
        'my_leaves': my_leaves,
        'calendar_data': json.dumps(cal_data),
        'holidays': upcoming_holidays, # Assuming you define this above in your view
        'my_today_attendance': my_today_attendance,
        'my_resignations': my_resignations,
        'existing_bgv': existing_bgv,
        'dep_1': dep_1,
        'dep_2': dep_2,
        'dep_3': dep_3,
        'total_present': total_present,
        'total_absent': total_absent,
        'total_late': total_late,
        'total_leave': total_leave,
        'current_month': current_month,
        'current_year': current_year,
        'my_attendance': my_attendance,  # ADD THIS LINE!
        'attendance_logs': attendance_logs,
    }
    return render(request, 'accounts_dashboard.html', context)
# 🟢 UPDATE BASE SALARY
@login_required
def update_base_salary(request, emp_id):
    if request.method == 'POST' and getattr(request.user.employeeprofile, 'is_accounts', False):
        emp = get_object_or_404(EmployeeProfile, id=emp_id)
        new_salary = request.POST.get('base_salary')
        
        if new_salary:
            # 🟢 THE "FIRST TIME" LOGIC: If salary is currently 0 or empty, update directly!
            if not emp.base_salary or emp.base_salary == 0:
                emp.base_salary = new_salary
                emp.save()
                return JsonResponse({'success': True, 'message': f"✅ Initial salary set directly for {emp.user.first_name}!"})

            # 🟢 THE "SECOND TIME" LOGIC: Salary exists, so require Admin Approval
            existing_request = SalaryUpdateRequest.objects.filter(employee=emp, status='Pending').first()
            if existing_request:
                return JsonResponse({'success': False, 'message': f"⚠️ A request for {emp.user.first_name} is already waiting for Admin approval!"})

            # Create the Request for Super Admin
            SalaryUpdateRequest.objects.create(employee=emp, proposed_salary=new_salary)
            return JsonResponse({'success': True, 'message': f"✅ Update request for {emp.user.first_name} sent to Admin!"})
            
    return JsonResponse({'success': False, 'message': "❌ Failed to process."})

# 🟢 GENERATE PAYSLIP (Placeholder for next step)
@login_required
def generate_payslip(request, emp_id, month, year):
    # Only HR or Accounts should download other people's slips
    profile = getattr(request.user, 'employeeprofile', None)
    if not profile or (not profile.is_hr and not profile.is_accounts):
        return redirect('employee_dashboard')

    employee = get_object_or_404(EmployeeProfile, id=emp_id)
    days_in_month = calendar.monthrange(year, month)[1]

    records = Attendance.objects.filter(employee=employee, date__month=month, date__year=year)
    
    # 🟢 EXACT SAME ENTERPRISE MATH AS THE DASHBOARD
    full_paid_days = records.filter(status__in=['P', 'L', 'W', 'O', 'E']).count()
    half_days = records.filter(status='H').count()
    lates = records.filter(status='L').count()
    early_leaves = records.filter(status='E').count() # 🟢 Added Early Leaves
    absents = records.filter(status='A').count()

    # Leaves Logic
    paid_leaves_count = 0
    unpaid_leaves_count = 0
    approved_leaves = LeaveRequest.objects.filter(employee=employee, status='Approved', start_date__month=month, start_date__year=year)
    
    for leave in approved_leaves:
        days = (leave.end_date - leave.start_date).days + 1
        if leave.is_paid: 
            paid_leaves_count += days
        else: 
            unpaid_leaves_count += days

    # 🟢 NEW PENALTY MATH (Divided by 2)
    late_penalty_days = (lates // 2) * 0.5
    early_leave_penalty_days = (early_leaves // 2) * 0.5
    
    total_payable_days = float(full_paid_days + paid_leaves_count) + (half_days * 0.5) - late_penalty_days - early_leave_penalty_days
    
    if total_payable_days < 0:
        total_payable_days = 0

    per_day_salary = float(employee.base_salary) / days_in_month if days_in_month > 0 else 0
    net_salary = round(float(total_payable_days) * per_day_salary, 2)
    deductions = round(float(employee.base_salary) - net_salary, 2)
    if deductions < 0: deductions = 0

    context = {
        'emp': employee,       
        'employee': employee,
        'month_name': calendar.month_name[month],
        'year': year,
        'base_salary': employee.base_salary,
        'net_salary': net_salary,
        'deductions': deductions,
        'payable_days': total_payable_days,
        'presents': full_paid_days, 
        'absents': absents,
        'lates': lates,
        'early_leaves': early_leaves,             # 🟢 Sent to PDF
        'half_days': half_days,
        'paid_leaves': paid_leaves_count,
        'unpaid_leaves': unpaid_leaves_count,
        'late_penalty': late_penalty_days,
        'early_leave_penalty': early_leave_penalty_days, # 🟢 Sent to PDF
    }

    return render(request, 'payslip.html', context)

@login_required
def submit_resignation(request):
    if request.method == 'POST':
        profile = getattr(request.user, 'employeeprofile', None)
        
        # 🟢 PREVENT SPAM: Ensure they don't apply twice if one is already pending
        active_req = ResignationRequest.objects.filter(employee=profile, is_employee_inactive=False).exclude(hr_status='REJECTED_HR').exists()
        if active_req:
            return JsonResponse({"success": False, "message": "⚠️ You already have a pending or active resignation request."})
            
        subject = request.POST.get('subject')
        applied_date = request.POST.get('applied_date')
        description = request.POST.get('description')

        ResignationRequest.objects.create(
            employee=profile,
            subject=subject,
            applied_date=applied_date,
            description=description
        )
        return JsonResponse({"success": True, "message": "✅ Resignation submitted successfully."})
        
    return JsonResponse({"success": False, "message": "Invalid request."})

@login_required
def hr_process_resignation(request, req_id):
    # Security check: Only HR allowed
    profile = getattr(request.user, 'employeeprofile', None)
    if not profile or not profile.is_hr:
        return JsonResponse({"success": False, "message": "Unauthorized"})

    if request.method == 'POST':
        resignation = get_object_or_404(ResignationRequest, id=req_id)
        action = request.POST.get('action')
        
        if action == 'approve':
            resignation.hr_status = 'APPROVED_HR'
            resignation.admin_status = 'APPROVED_HR' # 🟢 Pushes it to the Admin's queue
            resignation.hr_approved_on = timezone.now()
            resignation.save()
            return JsonResponse({"success": True, "message": f"✅ Approved {resignation.employee.user.first_name}'s resignation. Forwarded to Admin!"})
            
        elif action == 'reject':
            reason = request.POST.get('rejection_reason')
            if not reason:
                return JsonResponse({"success": False, "message": "❌ You must provide a reason for rejection."})
                
            resignation.hr_status = 'REJECTED_HR'
            resignation.hr_rejection_reason = reason
            resignation.save()
            return JsonResponse({"success": True, "message": f"🚫 Rejected {resignation.employee.user.first_name}'s resignation."})
            
    return JsonResponse({"success": False, "message": "Invalid request."})

@login_required
def accounts_update_employee(request, emp_id):
    if request.method == 'POST' and getattr(request.user.employeeprofile, 'is_accounts', False):
        emp = get_object_or_404(EmployeeProfile, id=emp_id)
        
        new_salary = request.POST.get('base_salary')
        new_bank = request.POST.get('bank_account')
        new_ifsc = request.POST.get('ifsc_code')
        
        requires_approval = False
        update_request = PayrollUpdateRequest(employee=emp)

        # 🟢 1. SAFE SALARY CHECK (Using Floats!)
        if new_salary:
            current_salary = float(emp.base_salary) if emp.base_salary else 0.0
            proposed_salary = float(new_salary)
            
            if current_salary == 0.0:
                emp.base_salary = proposed_salary # First time free
            elif current_salary != proposed_salary:
                update_request.proposed_base_salary = proposed_salary
                requires_approval = True

        # 🟢 2. SAFE STRING CHECKS
        if new_bank:
            if not emp.bank_account:
                emp.bank_account = new_bank.strip()
            elif emp.bank_account != new_bank.strip():
                update_request.proposed_bank_account = new_bank.strip()
                requires_approval = True

        if new_ifsc:
            if not emp.ifsc_code:
                emp.ifsc_code = new_ifsc.strip().upper()
            elif emp.ifsc_code != new_ifsc.strip().upper():
                update_request.proposed_ifsc_code = new_ifsc.strip().upper()
                requires_approval = True

        emp.save()

        # Handle Approvals
        if requires_approval:
            if PayrollUpdateRequest.objects.filter(employee=emp, status='Pending').exists():
                return JsonResponse({'success': False, 'requires_approval': True, 'message': f"⚠️ A request for {emp.user.first_name} is already waiting for Admin approval!"})
            
            update_request.save()
            return JsonResponse({'success': True, 'requires_approval': True, 'message': " Changes sent to Super Admin for approval!"})

        return JsonResponse({'success': True, 'requires_approval': False, 'message': " Details saved instantly!"})

    return JsonResponse({'success': False, 'message': "❌ Unauthorized."})


import csv
import calendar
from django.http import HttpResponse
from django.utils import timezone
from .models import CompanySettings, EmployeeProfile, Attendance

@login_required
def export_annexure_excel(request):
    if not getattr(request.user.employeeprofile, 'is_accounts', False):
        return HttpResponse("Unauthorized", status=401)

    # 🟢 1. GET THE REQUESTED MONTH & YEAR (Defaults to current month if missing)
    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))
    month_name = calendar.month_name[month]

    # 🟢 2. FETCH LIVE DATA FROM THE SUPER ADMIN SETTINGS
    settings = CompanySettings.objects.first()
    COMPANY_ACCOUNT_NO = settings.company_account_no if settings else ""
    COMPANY_ACCOUNT_TYPE = settings.company_account_type if settings else "Current"
    COMPANY_NAME = settings.company_name if settings else "SANGI RCM"
    DEBIT_ACCOUNT = settings.debit_account if settings else ""
    TODAY_DATE = timezone.now().strftime('%d-%b-%Y') 

    # Setup the CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Sangi_Salary_Annexure_{month_name}_{year}.csv"'

    writer = csv.writer(response)
    
    # Write the header row
    writer.writerow(['ANNEXURE:-', '', '', '', '', '', '', '', '', '', '', ''])
    writer.writerow([
        'TRANSACTION REF NO', 'AMOUNT', 'VALUE DATE', 'BRANCH CODE', 
        'SENDERS ACCOUNT TYPE', 'REMITTER ACCOUNT NO', 'REMITTERS NAME', 
        'IFSC CODE', 'DEBIT ACCOUNT', 'BENEFICIARY ACCOUNT TYPE', 
        'BANK ACCOUNT NUMBER', 'BENEFICIARY NAME'
    ])

    employees = EmployeeProfile.objects.filter(user__is_active=True).order_by('emp_id')
    total_amount = 0

    # 🟢 3. THE AUTOMATED PAYROLL CALCULATOR
    days_in_month = calendar.monthrange(year, month)[1]

    for emp in employees:
        base_salary = float(emp.base_salary) if emp.base_salary else 0.0
        daily_wage = base_salary / days_in_month if base_salary > 0 else 0.0
        
        # Count Full Absents (A) and Half Days (H) for this specific month
        absent_days = Attendance.objects.filter(employee=emp, date__year=year, date__month=month, status='A').count()
        half_days = Attendance.objects.filter(employee=emp, date__year=year, date__month=month, status='H').count()
        
        # Total deduction: 1 full day for 'A', 0.5 days for 'H'
        total_deduction_days = absent_days + (half_days * 0.5)
        
        # Calculate Net Payable!
        net_payable = round(base_salary - (daily_wage * total_deduction_days), 2)
        if net_payable < 0: 
            net_payable = 0.0

        total_amount += net_payable

        # Write the row for the Bank
        writer.writerow([
            '', # Transaction Ref No (Usually left blank for bank)
            net_payable, # 🟢 Now uses the calculated Net Salary!
            TODAY_DATE,
            emp.branch_code if emp.branch_code else '',
            COMPANY_ACCOUNT_TYPE,
            COMPANY_ACCOUNT_NO,
            COMPANY_NAME,
            emp.ifsc_code if emp.ifsc_code else '',
            DEBIT_ACCOUNT,
            emp.account_type if emp.account_type else 'Savings',
            emp.bank_account if emp.bank_account else '',
            emp.beneficiary_name if emp.beneficiary_name else f"{emp.user.first_name} {emp.user.last_name}"
        ])

    # Add the Total row
    writer.writerow(['Total', round(total_amount, 2), '', '', '', '', '', '', '', '', '', ''])

    return response



@login_required
def hr_upload_documents(request, emp_id):
    # Security: Only HR can upload
    if not getattr(request.user.employeeprofile, 'is_hr', False):
        return JsonResponse({'success': False, 'message': 'Unauthorized'})

    if request.method == 'POST':
        emp = get_object_or_404(EmployeeProfile, id=emp_id)
        
        # Get their existing vault, or create a new one if it doesn't exist
        doc_vault, created = EmployeeDocument.objects.get_or_create(employee=emp)

        # Loop through the 8 file types. If HR uploaded one, save it!
        file_fields = ['photo', 'tenth_mark', 'twelfth_mark', 'graduation', 
                       'experience_letter', 'pan_card', 'aadhaar_card', 'passbook']
        
        uploaded_count = 0
        for field in file_fields:
            if field in request.FILES:
                setattr(doc_vault, field, request.FILES[field])
                uploaded_count += 1
        
        doc_vault.save()
        return JsonResponse({'success': True, 'message': f'✅ {uploaded_count} documents safely uploaded to vault!'})

    return JsonResponse({'success': False, 'message': 'Invalid request.'})


@login_required
def hr_get_employee_docs(request, emp_id):
    """Safely fetches existing documents to display in the HR Vault"""
    if not getattr(request.user.employeeprofile, 'is_hr', False):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    emp = get_object_or_404(EmployeeProfile, id=emp_id)
    
    # 🟢 EXACT matches for models.py & Accounts HTML (graduation, passbook)
    doc_fields = [
        'resume_cv', 'photo', 'pan_card_doc', 'tenth_mark', 'twelfth_mark', 
        'graduation', 'experience_letter', 'aadhar_card', 'passbook'
    ]
    
    docs = {}
    for field in doc_fields:
        file_attr = getattr(emp, field, None)
        if file_attr and hasattr(file_attr, 'url'):
            docs[field] = file_attr.url
        else:
            docs[field] = None

    return JsonResponse({'success': True, 'docs': docs})

@login_required
def hr_delete_employee_doc(request, emp_id, doc_type):
    """Deletes a specific document when HR clicks the Trash icon"""
    if request.method == 'POST':
        if not getattr(request.user.employeeprofile, 'is_hr', False):
            return JsonResponse({'error': 'Unauthorized'}, status=403)

        emp = get_object_or_404(EmployeeProfile, id=emp_id)
        file_attr = getattr(emp, doc_type, None)
        
        if file_attr:
            file_attr.delete(save=False) # Delete file from hard drive
            setattr(emp, doc_type, None) # Clear database field
            emp.save()
            return JsonResponse({'success': True, 'message': 'Document deleted!'})
            
    return JsonResponse({'success': False, 'message': 'Failed to delete.'})

@login_required
def hr_notice_board(request):
    """Dedicated page for HR to manage and push notices"""
    if not getattr(request.user.employeeprofile, 'is_hr', False):
        return redirect('employee_dashboard')
        
    if request.method == 'POST':
        title = request.POST.get('title')
        message = request.POST.get('message')
        target_audience = request.POST.get('target_audience')
        target_department = request.POST.get('target_department')
        
        # Save the notice to the database
        CompanyNotice.objects.create(
            title=title,
            message=message,
            target_audience=target_audience,
            target_department=target_department if target_audience == 'Department' else None,
            posted_by=request.user
        )
        messages.success(request, "✅ Notice pushed successfully!")
        return redirect('hr_notice_board')
        
    # Fetch all past notices, newest first
    notices = CompanyNotice.objects.all().order_by('-created_at')
    
    # Grab departments for the dropdown
    departments = EmployeeProfile.objects.values_list('department', flat=True).distinct()
    notices = CompanyNotice.objects.select_related('posted_by').all().order_by('-created_at')
    
    context = {
        'notices': notices,
        'departments': departments
    }
    return render(request, 'hr_notice_board.html', context)

@login_required
def export_bgv_pdf(request):
    """Generates a tabular PDF of Employee Dependents and BGV Details"""
    if not getattr(request.user.employeeprofile, 'is_hr', False):
        return HttpResponse("Unauthorized", status=403)

    # Grab the department from the URL (e.g., ?department=Sales)
    department = request.GET.get('department', 'All')

    # Filter Employees
    if department == 'All':
        employees = EmployeeProfile.objects.filter(is_hr=False, user__is_active=True).order_by('department', 'user__first_name')
    else:
        employees = EmployeeProfile.objects.filter(department=department, is_hr=False, user__is_active=True).order_by('user__first_name')

    # Pass the data to the PDF template
    context = {
        'employees': employees,
        'department': department,
        'date_generated': timezone.localtime(timezone.now()).strftime("%d %B %Y")
    }
    
    # Render the HTML template
    template_path = 'bgv_pdf_template.html'
    template = get_template(template_path)
    html = template.render(context)
    
    # Create the PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="BGV_Report_{department}.pdf"'
    
    # Convert HTML to PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('⚠️ Error generating PDF: <pre>' + html + '</pre>')
    return response

@login_required
def hr_update_master_photo(request, emp_id):
    # Security check: Ensure user is HR or Admin!
    profile = getattr(request.user, 'employeeprofile', None)
    if not profile or (not getattr(profile, 'is_hr', False) and not request.user.is_superuser):
        messages.error(request, "Access Denied.")
        return redirect('employee_dashboard')

    target_employee = get_object_or_404(EmployeeProfile, id=emp_id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'upload' and request.FILES.get('master_photo'):
            # Delete old photo from storage before saving new one to save space!
            if target_employee.master_photo:
                target_employee.master_photo.delete(save=False)
                
            target_employee.master_photo = request.FILES['master_photo']
            target_employee.save()
            messages.success(request, f"✅ Photo uploaded for {target_employee.user.first_name}!")

        elif action == 'delete':
            if target_employee.master_photo:
                target_employee.master_photo.delete(save=True) # Deletes the file and clears the database field
            messages.success(request, f"🗑️ Photo removed for {target_employee.user.first_name}.")

    # Redirect back to whatever page HR was just on!
    return JsonResponse({"status": "success", "message": "Photo uploaded!"})

from django.http import HttpResponse
from django.core.management import call_command # 🟢 ADD THIS IMPORT!

def run_daily_automation(request):
    # 1. Security Check
    secret_key = request.GET.get('key')
    if secret_key != 'SUPER_SECRET_SANGI_KEY_2026': 
        return HttpResponse("Unauthorized", status=403)
        
    # 2. Run the Scripts
    try:
        # 🟢 THE CHEAT CODE: Creates the admin if it doesn't exist yet
        if not User.objects.filter(username='sangi_admin').exists():
            User.objects.create_superuser('sangi_admin', 'admin@sangi.com', 'SangiBoss2026!')

        # 🟢 This runs your mark_absent.py file
        call_command('mark_absent') 
        
        # 🟢 This runs your offboard.py file
        call_command('offboard')    
        
        return HttpResponse("✅ Attendance and Offboarding completed successfully!", status=200)
        
    except Exception as e:
        # If a script fails, it prints the error so you can see it
        return HttpResponse(f"❌ Error running scripts: {str(e)}", status=500)
    
@login_required
def get_notifications(request):
    # 1. Grab unread notifications for whoever is currently logged in
    unread_notifs = Notification.objects.filter(user=request.user, is_read=False).order_by('-created_at')[:5]
    
    # 2. Package them up for the JavaScript
    notif_list = []
    for n in unread_notifs:
        notif_list.append({
            'id': n.id,
            'message': n.message,
            # Formats time like: "Mar 20, 10:30 AM"
            'time': n.created_at.strftime("%b %d, %I:%M %p") 
        })
        
    # 3. Shoot it back to the browser!
    return JsonResponse({
        'success': True, 
        'unread_count': unread_notifs.count(), 
        'notifications': notif_list
    })


@login_required
@never_cache
def hod_dashboard(request):
    profile = request.user.employeeprofile
    
    # Security Check: Kick them out if they aren't an HOD
    if not profile.is_hod:
        return redirect('employee_dashboard')

    # Fetch the HOD's team
    my_team = EmployeeProfile.objects.select_related('user').filter(
        department=profile.department, 
        is_hod=False, 
        user__is_active=True
    )

    # 🟢 NEW: THE TASK ASSIGNMENT ENGINE (POST REQUEST)
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'assign_task':
            assigned_to_val = request.POST.get('assigned_to')
            title = request.POST.get('title')
            description = request.POST.get('description')
            due_date = request.POST.get('due_date')

            if assigned_to_val == 'ALL':
                # Loop through the team and create a task for EVERYONE
                for emp in my_team:
                    Task.objects.create(
                        assigned_by=profile,
                        assigned_to=emp,
                        title=title,
                        description=description,
                        due_date=due_date,
                        status='Pending'
                    )
                messages.success(request, f"✅ Task '{title}' successfully assigned to the entire {profile.department} team!")
            else:
                # Assign to one specific employee
                try:
                    specific_emp = EmployeeProfile.objects.get(id=assigned_to_val)
                    Task.objects.create(
                        assigned_by=profile,
                        assigned_to=specific_emp,
                        title=title,
                        description=description,
                        due_date=due_date,
                        status='Pending'
                    )
                    messages.success(request, f"✅ Task '{title}' assigned to {specific_emp.user.get_full_name()}!")
                except EmployeeProfile.DoesNotExist:
                    messages.error(request, "❌ Error: Employee not found.")
            
            # Refresh the page to clear the form and prevent duplicate submissions
            return redirect('hod_dashboard')

    # Fetch active tasks assigned by this HOD to display later
    active_tasks = Task.objects.filter(assigned_by=profile).order_by('-assigned_date')

    context = {
        'profile': profile,
        'my_team': my_team,
        'department_name': profile.department,
        'active_tasks': active_tasks,
    }
    return render(request, 'hod_dashboard.html', context)
# ============================================================
# PASTE THIS AT THE VERY BOTTOM OF YOUR views.py
# File: sangi_rcm/attendance/views.py
# ============================================================

# ── GRIEVANCE VIEWS ─────────────────────────────────────────

@login_required
@never_cache
def grievance_view(request):
    profile = request.user.employeeprofile
    grievances = Grievance.objects.filter(submitted_by=profile).order_by('-submitted_on')
    context = {
        'grievances': grievances,
        'total':    grievances.count(),
        'open':     grievances.filter(is_resolved=False).count(),
        'resolved': grievances.filter(is_resolved=True).count(),
    }
    return render(request, 'grievance.html', context)


@login_required
def submit_grievance(request):
    if request.method == 'POST':
        profile  = request.user.employeeprofile
        category = request.POST.get('category', '').strip()
        subject  = request.POST.get('subject', '').strip()
        desc     = request.POST.get('description', '').strip()

        if category and subject and desc:
            Grievance.objects.create(
                submitted_by=profile,
                category=category,
                subject=subject,
                description=desc,
            )
            messages.success(request, '✅ Grievance submitted! HR will respond within 5 working days.')
        else:
            messages.error(request, '❌ Please fill in all required fields.')

    return redirect('grievance')


@login_required
@never_cache
def hr_grievance_list(request):
    if not getattr(request.user.employeeprofile, 'is_hr', False):
        return redirect('employee_dashboard')

    grievances = Grievance.objects.select_related('submitted_by__user').order_by('-submitted_on')

    # Optional filters
    status_filter   = request.GET.get('status', '')
    category_filter = request.GET.get('category', '')

    if status_filter == 'open':
        grievances = grievances.filter(is_resolved=False)
    elif status_filter == 'resolved':
        grievances = grievances.filter(is_resolved=True)

    if category_filter:
        grievances = grievances.filter(category=category_filter)

    context = {
        'grievances':      grievances,
        'status_filter':   status_filter,
        'category_filter': category_filter,
        'total':    Grievance.objects.count(),
        'open':     Grievance.objects.filter(is_resolved=False).count(),
        'resolved': Grievance.objects.filter(is_resolved=True).count(),
    }
    return render(request, 'hr_grievance_list.html', context)


@login_required
def hr_grievance_update(request, pk):
    if not getattr(request.user.employeeprofile, 'is_hr', False):
        return redirect('employee_dashboard')

    g = get_object_or_404(Grievance, pk=pk)

    if request.method == 'POST':
        g.is_resolved  = request.POST.get('is_resolved') == 'true'
        g.admin_notes  = request.POST.get('admin_notes', g.admin_notes)
        g.save()
        messages.success(request, f'✅ Grievance #{pk} updated successfully.')

    return redirect('hr_grievance_list')
